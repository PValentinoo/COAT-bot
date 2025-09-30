import argparse
import json
import os
import sys
import time
import random
from collections import OrderedDict, defaultdict
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional, Iterable

import requests
from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet


CACHE_FILENAME = ".formula_explanations_cache.json"
DEFAULT_MODEL = "gpt-4o-mini"
MAX_CACHE_ENTRIES = 1000


@dataclass
class FormulaInfo:
    formula_text: str
    locations: List[Tuple[str, str]]  # list of (sheet_name, a1_address)


def load_cache(cache_path: str) -> OrderedDict:
    """Load cache as an OrderedDict for simple LRU behavior.

    Structure on disk: {"items": {formula: {"explanation": str, "last_used": float, "uses": int}}, "version": 1}
    We restore order by last_used descending so the first items are most recently used.
    """
    if not os.path.exists(cache_path):
        return OrderedDict()
    try:
        with open(cache_path, "r", encoding="utf-8") as f:
            data = json.load(f)
        items = data.get("items", {})
        # Sort by last_used descending (most recent first)
        sorted_items = sorted(items.items(), key=lambda kv: kv[1].get("last_used", 0.0), reverse=True)
        od: OrderedDict = OrderedDict()
        for k, v in sorted_items:
            od[k] = v
        return od
    except Exception:
        # Corrupt cache — start fresh
        return OrderedDict()


def save_cache(cache_path: str, cache: OrderedDict) -> None:
    """Persist cache. Trim to MAX_CACHE_ENTRIES using LRU logic (keep most recently used)."""
    # Ensure size limit
    trimmed_items: List[Tuple[str, dict]] = list(cache.items())[:MAX_CACHE_ENTRIES]
    data = {
        "version": 1,
        "items": {k: v for k, v in trimmed_items},
        "saved_at": time.time(),
        "size": len(trimmed_items),
    }
    with open(cache_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)


def is_formula(cell_value: object) -> bool:
    return isinstance(cell_value, str) and cell_value.startswith("=")


def a1_address(cell) -> str:
    return f"{cell.column_letter}{cell.row}"


def collect_formulas(wb: Workbook) -> Dict[str, List[Tuple[str, str]]]:
    """Traverse all sheets and collect formulas.

    Returns mapping: formula_text -> list of (sheet_name, A1_address)
    """
    mapping: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    total_cells_scanned = 0
    for ws in wb.worksheets:  # type: Worksheet
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                total_cells_scanned += 1
                val = cell.value
                if is_formula(val):
                    mapping[val].append((ws.title, a1_address(cell)))
    print(f"Scannet celler i alt: {total_cells_scanned}")
    return mapping


def format_example_locations(locations: List[Tuple[str, str]], max_examples: int) -> str:
    examples = [f"{sheet}!{addr}" for sheet, addr in locations[:max_examples]]
    return ", ".join(examples)


def ensure_api_key_or_exit(dry_run: bool) -> str:
    if dry_run:
        return ""
    # Load from .env if present (simple loader: KEY=VALUE lines)
    env_path_candidates = [
        os.path.join(os.getcwd(), ".env"),
        os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env"),
    ]
    for env_path in env_path_candidates:
        if os.path.exists(env_path):
            try:
                with open(env_path, "r", encoding="utf-8") as f:
                    for line in f:
                        line = line.strip()
                        if not line or line.startswith("#"):
                            continue
                        if "=" in line:
                            key, value = line.split("=", 1)
                            key = key.strip()
                            value = value.strip().strip('"').strip("'")
                            if key and key not in os.environ:
                                os.environ[key] = value
            except Exception:
                pass
    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        print("FEJL: Miljøvariablen OPENAI_API_KEY mangler. Sæt nøglen og prøv igen.")
        sys.exit(1)
    return api_key


def call_openai_with_backoff(model: str, prompt_formula: str, api_key: str, temperature: float = 0.2, max_retries: int = 5) -> str:
    """Call OpenAI Chat Completions API with exponential backoff on 429/5xx."""
    url = "https://api.openai.com/v1/chat/completions"
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json",
    }
    system_prompt = (
        "Du er en dansk Excel-ekspert. Forklar præcist, pædagogisk og uden overflødigt sprog. "
        "Brug semikolon i funktions-eksempler."
    )
    user_prompt = (
        "Giv en detaljeret dansk forklaring af følgende Excel-formel. Beskriv hvad den gør, de centrale "
        "funktioner (IF, IFERROR, DATEDIF, ROUNDUP, PMT, XLOOKUP, SWITCH m.fl.), og hvordan absolutte/relative "
        f"referencer påvirker beregningen. Formel: {prompt_formula}"
    )

    body = {
        "model": model,
        "temperature": temperature,
        "messages": [
            {"role": "system", "content": system_prompt},
            {"role": "user", "content": user_prompt},
        ],
    }

    delay = 1.0
    for attempt in range(1, max_retries + 1):
        try:
            resp = requests.post(url, headers=headers, json=body, timeout=60)
            if resp.status_code == 200:
                data = resp.json()
                content = data["choices"][0]["message"]["content"].strip()
                return content
            elif resp.status_code in (429, 500, 502, 503, 504):
                # Backoff and retry
                if attempt == max_retries:
                    raise RuntimeError(f"OpenAI API fejlede efter {max_retries} forsøg: {resp.status_code} {resp.text}")
                sleep_time = delay * (2 ** (attempt - 1)) + random.uniform(0, 0.5)
                time.sleep(sleep_time)
                continue
            else:
                # Non-retryable
                raise RuntimeError(f"OpenAI API fejl: {resp.status_code} {resp.text}")
        except requests.RequestException as e:
            if attempt == max_retries:
                raise RuntimeError(f"Netværksfejl efter {max_retries} forsøg: {e}") from e
            sleep_time = delay * (2 ** (attempt - 1)) + random.uniform(0, 0.5)
            time.sleep(sleep_time)
            continue


def explain_formulas(
    mapping: Dict[str, List[Tuple[str, str]]],
    model: str,
    dry_run: bool,
    cache: OrderedDict,
    api_key: str,
) -> Tuple[Dict[str, str], int]:
    """Return mapping formula->explanation and number of cache hits."""
    explanations: Dict[str, str] = {}
    cache_hits = 0
    to_explain = sorted(mapping.keys())
    for formula in to_explain:
        cached = cache.get(formula)
        if cached and isinstance(cached, dict) and cached.get("explanation"):
            explanations[formula] = cached["explanation"]
            cached["last_used"] = time.time()
            cached["uses"] = int(cached.get("uses", 0)) + 1
            # Move to front (most recent) by reinserting at beginning
            cache.move_to_end(formula, last=False)
            cache_hits += 1
            continue

        if dry_run:
            explanation = "[Tørkørsel] Placeholder-forklaring for denne formel."
        else:
            explanation = call_openai_with_backoff(model=model, prompt_formula=formula, api_key=api_key)

        explanations[formula] = explanation
        cache[formula] = {
            "explanation": explanation,
            "last_used": time.time(),
            "uses": 1,
        }
        # Most recent at front
        cache.move_to_end(formula, last=False)
    return explanations, cache_hits


def write_explanations_sheet(
    wb: Workbook,
    sheet_name: str,
    mapping: Dict[str, List[Tuple[str, str]]],
    explanations: Dict[str, str],
    max_examples: int,
) -> None:
    # Remove old sheet if exists
    if sheet_name in wb.sheetnames:
        ws_old = wb[sheet_name]
        wb.remove(ws_old)
    ws = wb.create_sheet(sheet_name)

    # Headers
    ws["A1"] = "Formel"
    ws["B1"] = "Forklaring (dansk)"
    ws["C1"] = "Forekomster"
    ws["D1"] = "Eksempler (ark!celle)"

    # Sort: by count desc, then formula asc
    rows = []
    for formula, locs in mapping.items():
        rows.append((len(locs), formula))
    rows.sort(key=lambda x: (-x[0], x[1]))

    r = 2
    for count, formula in rows:
        ws.cell(row=r, column=1, value=formula)
        ws.cell(row=r, column=2, value=explanations.get(formula, ""))
        ws.cell(row=r, column=3, value=count)
        ws.cell(row=r, column=4, value=format_example_locations(mapping[formula], max_examples))
        r += 1


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Find unikke Excel-formler og generér danske forklaringer.")
    parser.add_argument("--path", required=True, help="Sti til .xlsx arbejdsbogen")
    parser.add_argument("--sheet", default="Forklaringer", help="Navn på forklarings-arket (default: Forklaringer)")
    parser.add_argument("--model", default=DEFAULT_MODEL, help=f"LLM modelnavn (default: {DEFAULT_MODEL})")
    parser.add_argument("--max-examples", type=int, default=10, help="Maks antal eksempel-lokationer per formel")
    parser.add_argument("--dry-run", action="store_true", help="Spring LLM over og brug placeholder-forklaringer")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    xlsx_path: str = args.path
    sheet_name: str = args.sheet
    model: str = args.model
    max_examples: int = args.max_examples
    dry_run: bool = args.dry_run

    if not os.path.exists(xlsx_path):
        print(f"FEJL: Filen findes ikke: {xlsx_path}")
        sys.exit(1)

    api_key = ensure_api_key_or_exit(dry_run=dry_run)

    # Load workbook with formulas
    wb = load_workbook(filename=xlsx_path, data_only=False)

    # Collect formulas
    mapping = collect_formulas(wb)
    num_unique = len(mapping)
    num_cells_with_formulas = sum(len(v) for v in mapping.values())
    print(f"Unikke formler: {num_unique}")
    print(f"Formelceller i alt: {num_cells_with_formulas}")

    # Load cache
    cache_path = os.path.join(os.path.dirname(os.path.abspath(xlsx_path)) or os.getcwd(), CACHE_FILENAME)
    cache = load_cache(cache_path)

    # Determine which will hit LLM
    needs_llm = [f for f in mapping.keys() if f not in cache]
    print(f"Forventede LLM-kald: {0 if dry_run else len(needs_llm)} (tørkørsel: {dry_run})")

    # Explain
    explanations, cache_hits = explain_formulas(mapping, model, dry_run, cache, api_key)
    print(f"Cache-hits: {cache_hits}")

    # Write sheet
    write_explanations_sheet(wb, sheet_name, mapping, explanations, max_examples)

    # Save workbook (in-place)
    wb.save(xlsx_path)
    print(f"Fil gemt: {xlsx_path}")

    # Save cache
    save_cache(cache_path, cache)
    print(f"Cache opdateret: {cache_path}")


if __name__ == "__main__":
    main()


