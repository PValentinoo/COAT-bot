import argparse
import os
import re
import sys
from collections import defaultdict
from typing import Dict, List, Tuple, Optional

try:
    import tkinter as tk
    from tkinter import filedialog
    TK_AVAILABLE = True
except Exception:
    TK_AVAILABLE = False

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo


# Set this to True or False to control default similarity grouping when running via GUI
DEFAULT_SIMILARITY: bool = True


def is_formula(value: object) -> bool:
    return isinstance(value, str) and value.startswith("=")


def a1(cell) -> str:
    return f"{cell.column_letter}{cell.row}"


def pick_file_gui() -> str:
    if not TK_AVAILABLE:
        return ""
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    path = filedialog.askopenfilename(
        title="Vælg en Excel-fil",
        filetypes=[
            ("Excel-filer (.xlsx, .xlsm, .xltx, .xltm)", "*.xlsx *.xlsm *.xltx *.xltm"),
            ("Alle filer", "*.*"),
        ],
    )
    root.update()
    root.destroy()
    return path or ""


def pick_destination_dir_gui() -> str:
    if not TK_AVAILABLE:
        return ""
    root = tk.Tk()
    root.withdraw()
    root.wm_attributes('-topmost', 1)
    path = filedialog.askdirectory(title="Vælg destinationsmappe til rapporter")
    root.update()
    root.destroy()
    return path or ""


def collect_formulas(xlsx_path: str) -> Dict[str, List[Tuple[str, str]]]:
    wb = load_workbook(filename=xlsx_path, data_only=False)
    mapping: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    total_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                total_cells += 1
                v = cell.value
                if is_formula(v):
                    mapping[v].append((ws.title, a1(cell)))
    return mapping


def collect_text_cells(xlsx_path: str) -> Dict[str, List[Tuple[str, str]]]:
    """Collect non-formula text cell contents across the workbook.

    Returns mapping: text_value -> list of (sheet_name, A1_address)
    """
    wb = load_workbook(filename=xlsx_path, data_only=False)
    mapping: Dict[str, List[Tuple[str, str]]] = defaultdict(list)
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=False):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and not v.startswith("="):
                    mapping[v].append((ws.title, a1(cell)))
    return mapping


def summarize(mapping: Dict[str, List[Tuple[str, str]]], max_examples: int) -> str:
    lines: List[str] = []
    total_formula_cells = sum(len(locs) for locs in mapping.values())
    lines.append(f"Formelceller i alt: {total_formula_cells}")
    lines.append(f"Unikke formler: {len(mapping)}")
    lines.append("")
    rows = [
        (len(locs), formula, locs)
        for formula, locs in mapping.items()
    ]
    rows.sort(key=lambda x: (-x[0], x[1]))
    for count, formula, locs in rows:
        examples = ", ".join([f"{s}!{a}" for s, a in locs[:max_examples]])
        lines.append(f"- Forekomster: {count} | Længde: {len(formula)} | Formel: {formula}")
        lines.append(f"  Eksempler ({min(max_examples, len(locs))}): {examples}")
    return "\n".join(lines)


# Normalization helpers for similarity groups
ABS_REF_RE = re.compile(r"\$")
NUMBER_RE = re.compile(r"(?<![A-Za-z_])[+-]?(?:\d+\.\d+|\d+)(?:[eE][+-]?\d+)?")
STRING_RE = re.compile(r'"[^"]*"')
SHEET_REF_RE = re.compile(r"(?:(?:'[^']+'|[A-Za-z0-9_]+))!")
WHITESPACE_RE = re.compile(r"\s+")
# Replace row numbers in cell refs (e.g., a147 -> a#). Handles 1-3 letter columns.
CELL_ROW_RE = re.compile(r"(?<![A-Za-z0-9_])([a-z]{1,3})\d+")

# Direct reference formula matcher: matches an entire formula that is only a single cell or range ref
DIRECT_REF_RE = re.compile(
    r"^=\+?\s*(?:(?:'[^']+'|[A-Za-z0-9_]+)!)?\$?[A-Za-z]{1,3}\$?\d+(?::\$?[A-Za-z]{1,3}\$?\d+)?\s*$",
    re.IGNORECASE,
)

# Regex to capture single cell references with optional sheet qualifier
CELL_REF_WITH_SHEET_RE = re.compile(
    r"(?:(?:'(?P<qsheet>[^']+)'|(?P<sheet>[A-Za-z0-9_]+))!)?\$?(?P<col>[A-Za-z]{1,3})\$?(?P<row>\d+)",
    re.IGNORECASE,
)


def render_formula_with_values(formula: str, default_sheet: str, wb_values) -> str:
    """Replace single cell references in the formula with their current values.

    - Uses wb_values (data_only=True) to get values.
    - Unqualified refs default to default_sheet.
    - Leaves ranges and unknown refs as-is.
    - Returns a plain string; caller may prefix with ' to prevent evaluation.
    """
    if wb_values is None or not default_sheet:
        return formula

    def _replace(match: re.Match) -> str:
        sheet = match.group('qsheet') or match.group('sheet') or default_sheet
        col = match.group('col')
        row = match.group('row')
        a1 = f"{col.upper()}{row}"
        # If this looks like part of a larger token (e.g., function name), keep as-is
        try:
            ws = wb_values[sheet]
            value = ws[a1].value
            if value is None:
                return ""  # empty string for None
            return str(value)
        except Exception:
            return match.group(0)

    # Quick exit: if no '=' assume not a formula string
    if not isinstance(formula, str) or not formula.startswith("="):
        return formula

    # Replace only single-cell refs; ranges like A1:B2 will be left untouched
    try:
        return CELL_REF_WITH_SHEET_RE.sub(_replace, formula)
    except Exception:
        return formula


def normalize_formula(formula: str) -> str:
    s = formula.strip().lower()
    # Normalize '=+...' to '=' and unify separators to comma
    if s.startswith("=+"):
        s = "=" + s[2:]
    s = s.replace(";", ",")
    s = ABS_REF_RE.sub("", s)
    s = SHEET_REF_RE.sub("", s)
    s = STRING_RE.sub('"…"', s)
    # Replace numbers not part of cell refs
    s = NUMBER_RE.sub("#", s)
    # Replace row numbers in cell references with '#'
    s = CELL_ROW_RE.sub(r"\1#", s)
    s = WHITESPACE_RE.sub(" ", s)
    return s


def summarize_similarity(mapping: Dict[str, List[Tuple[str, str]]], max_examples: int, top_variants: int = 5) -> str:
    groups: Dict[str, List[str]] = defaultdict(list)
    for formula in mapping.keys():
        # Skip trivial direct cell/range references from grouping
        if DIRECT_REF_RE.match(formula):
            continue
        groups[normalize_formula(formula)].append(formula)

    lines: List[str] = []
    lines.append("Lighedsoverblik (normaliserede grupper):")
    lines.append("")

    group_rows: List[Tuple[int, str]] = []
    for norm, originals in groups.items():
        total = sum(len(mapping[o]) for o in set(originals))
        group_rows.append((total, norm))
    group_rows.sort(key=lambda x: (-x[0], x[1]))

    for total, norm in group_rows:
        lines.append(f"=== Gruppe (forekomster: {total}) ===")
        lines.append(f"Normaliseret: {norm}")
        variant_counts = [(len(mapping[o]), o) for o in set(groups[norm])]
        variant_counts.sort(key=lambda x: (-x[0], x[1]))
        for count, original in variant_counts[:top_variants]:
            examples = ", ".join([f"{s}!{a}" for s, a in mapping[original][:max_examples]])
            lines.append(f"- {count}x | {original}")
            lines.append(f"  Eksempler: {examples}")
        lines.append("")
    return "\n".join(lines)


def export_to_excel(
    xlsx_input: str,
    mapping: Dict[str, List[Tuple[str, str]]],
    similarity: bool,
    max_examples: int,
    out_path: str,
) -> str:
    wb_out = Workbook()

    # Summary sheet
    ws_sum = wb_out.active
    ws_sum.title = "Summary"
    total_formula_cells = sum(len(locs) for locs in mapping.values())
    ws_sum["A1"] = "Input fil"
    ws_sum["B1"] = xlsx_input
    ws_sum["A2"] = "Formelceller i alt"
    ws_sum["B2"] = total_formula_cells
    ws_sum["A3"] = "Unikke formler"
    ws_sum["B3"] = len(mapping)
    ws_sum.freeze_panes = "A2"

    # Context sheet: unique non-formula text values with counts and examples
    try:
        texts_mapping = collect_text_cells(xlsx_input)
    except Exception:
        texts_mapping = {}
    ws_txt = wb_out.create_sheet("Context")
    ws_txt["A1"] = "Ark"
    ws_txt["B1"] = "Tekst"
    ws_txt["C1"] = "Forekomster"
    ws_txt["D1"] = "Længde"
    ws_txt["E1"] = "Eksempler (ark!celle)"
    # Build per-sheet rows so 'Ark' is accurate, sort Ark alphabetically
    txt_rows = []
    for text, locs in texts_mapping.items():
        per_sheet: Dict[str, List[str]] = defaultdict(list)
        for sheet, a1 in locs:
            per_sheet[sheet].append(a1)
        for sheet, addrs in per_sheet.items():
            examples = ", ".join([f"{sheet}!{a}" for a in addrs[:max_examples]])
            txt_rows.append((sheet, text, len(addrs), len(text), examples))
    txt_rows.sort(key=lambda x: (x[0], x[1]))
    r = 2
    for sheet, text, count, length, examples in txt_rows:
        ws_txt.cell(row=r, column=1, value=sheet)
        ws_txt.cell(row=r, column=2, value=text)
        ws_txt.cell(row=r, column=3, value=count)
        ws_txt.cell(row=r, column=4, value=length)
        ws_txt.cell(row=r, column=5, value=examples)
        r += 1
    ws_txt.freeze_panes = "A2"
    ws_txt.column_dimensions['A'].width = 18
    ws_txt.column_dimensions['B'].width = 80
    ws_txt.column_dimensions['C'].width = 14
    ws_txt.column_dimensions['D'].width = 10
    ws_txt.column_dimensions['E'].width = 50
    last_row_txt = max(2, ws_txt.max_row)
    table_txt = Table(displayName="ContextTable", ref=f"A1:E{last_row_txt}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table_txt.tableStyleInfo = style
    ws_txt.add_table(table_txt)

    # Unique formulas sheet (repurposed to Normalized summary with example value), placed after Context
    ws_uniq = wb_out.create_sheet("Unique Formulas")
    ws_uniq["A1"] = "Ark"
    ws_uniq["B1"] = "Forekomster (gruppe)"
    ws_uniq["C1"] = "Normaliseret"
    ws_uniq["D1"] = "Eksempel formel"
    ws_uniq["E1"] = "Eksempel lokation"

    # Build normalized groups (skip direct refs)
    groups: Dict[str, List[str]] = defaultdict(list)
    for formula in mapping.keys():
        if DIRECT_REF_RE.match(formula):
            continue
        groups[normalize_formula(formula)].append(formula)

    # Open a second workbook view with data_only=True to fetch stored result values
    try:
        wb_values = load_workbook(filename=xlsx_input, data_only=True)
    except Exception:
        wb_values = None

    # Prepare rows: sort by total occurrences desc
    group_rows: List[Tuple[int, str]] = []
    for norm, originals in groups.items():
        total = sum(len(mapping[o]) for o in set(originals))
        group_rows.append((total, norm))
    group_rows.sort(key=lambda x: (-x[0], x[1]))

    uniq_rows = []
    for total, norm in group_rows:
        variants = [(len(mapping[o]), o) for o in set(groups[norm])]
        variants.sort(key=lambda x: (-x[0], x[1]))
        example_formula = variants[0][1]
        example_loc = mapping[example_formula][0] if mapping[example_formula] else ("", "")
        example_addr = f"{example_loc[0]}!{example_loc[1]}" if example_loc[0] else ""
        default_sheet = example_loc[0] if example_loc[0] else ""
        rendered_with_values = render_formula_with_values(example_formula, default_sheet, wb_values)
        uniq_rows.append((example_loc[0], total, norm, rendered_with_values, example_addr))
    # Sort by Ark alphabetically
    uniq_rows.sort(key=lambda x: (x[0] or "" , x[2]))
    r = 2
    for sheet_name, total, norm, rendered_with_values, example_addr in uniq_rows:
        ws_uniq.cell(row=r, column=1, value=sheet_name)
        ws_uniq.cell(row=r, column=2, value=total)
        ws_uniq.cell(row=r, column=3, value=f"'{norm}")
        ws_uniq.cell(row=r, column=4, value=f"'{rendered_with_values}")
        ws_uniq.cell(row=r, column=5, value=example_addr)
        r += 1

    ws_uniq.freeze_panes = "A2"
    # Column widths for readability
    ws_uniq.column_dimensions['A'].width = 18
    ws_uniq.column_dimensions['B'].width = 18
    ws_uniq.column_dimensions['C'].width = 100
    ws_uniq.column_dimensions['D'].width = 100
    ws_uniq.column_dimensions['E'].width = 24
    # Add table with filters
    last_row = max(2, ws_uniq.max_row)
    table_uniq = Table(displayName="NormalizedSummary", ref=f"A1:E{last_row}")
    style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=True, showColumnStripes=False)
    table_uniq.tableStyleInfo = style
    ws_uniq.add_table(table_uniq)

    if similarity:
        # Raw data sheet (formerly Similarity Groups)
        ws_sim = wb_out.create_sheet("Raw data")
        ws_sim["A1"] = "Ark"
        ws_sim["B1"] = "Gruppe total"
        ws_sim["C1"] = "Normaliseret"
        ws_sim["D1"] = "Variant formel"
        ws_sim["E1"] = "Eksempler"

        # Build normalized groups
        groups: Dict[str, List[str]] = defaultdict(list)
        for formula in mapping.keys():
            # Skip trivial direct cell/range references from grouping
            if DIRECT_REF_RE.match(formula):
                continue
            groups[normalize_formula(formula)].append(formula)

        # Sort groups by total occurrences desc
        group_rows: List[Tuple[int, str]] = []
        for norm, originals in groups.items():
            total = sum(len(mapping[o]) for o in set(originals))
            group_rows.append((total, norm))
        group_rows.sort(key=lambda x: (-x[0], x[1]))

        # Collect rows then sort by Ark alphabetically
        raw_rows = []
        for total, norm in group_rows:
            variant_counts = [(len(mapping[o]), o) for o in set(groups[norm])]
            variant_counts.sort(key=lambda x: (-x[0], x[1]))
            for count, original in variant_counts:
                examples = ", ".join([f"{s}!{a}" for s, a in mapping[original][:max_examples]])
                first_sheet = mapping[original][0][0] if mapping[original] else ""
                raw_rows.append((first_sheet, total, norm, original, examples))
        raw_rows.sort(key=lambda x: (x[0] or "", x[2], x[3]))
        r = 2
        for sheet_name, total, norm, original, examples in raw_rows:
            ws_sim.cell(row=r, column=1, value=sheet_name)
            ws_sim.cell(row=r, column=2, value=total)
            ws_sim.cell(row=r, column=3, value=f"'{norm}")
            ws_sim.cell(row=r, column=4, value=f"'{original}")
            ws_sim.cell(row=r, column=5, value=examples)
            r += 1
        ws_sim.freeze_panes = "A2"
        ws_sim.column_dimensions['A'].width = 18
        ws_sim.column_dimensions['B'].width = 14
        ws_sim.column_dimensions['C'].width = 100
        ws_sim.column_dimensions['D'].width = 100
        ws_sim.column_dimensions['E'].width = 50
        last_row_sim = max(2, ws_sim.max_row)
        table_sim = Table(displayName="RawDataTable", ref=f"A1:E{last_row_sim}")
        table_sim.tableStyleInfo = style
        ws_sim.add_table(table_sim)

    # (Removed legacy 'Texts' tab; 'Context' tab already provides this information.)

    # Build aggregated strings for Summary sheet (for chatbot copy-paste)
    # Context aggregation: 25 longest unique texts
    try:
        all_texts = list(texts_mapping.keys())
    except Exception:
        all_texts = []
    all_texts.sort(key=lambda t: len(t or ""), reverse=True)
    top_texts = all_texts[:25]
    context_agg = "\n\n".join(t for t in top_texts if t)

    # Formulas aggregation: one line per normalized group
    formula_lines = []
    for sheet_name, total, norm, rendered_with_values, example_addr in uniq_rows:
        line = (
            f"Ark: {sheet_name} | Forekomster: {total} | Normaliseret: {norm} | "
            f"Eksempel: {rendered_with_values} | Lokation: {example_addr}"
        )
        formula_lines.append(line)
    formulas_agg = "\n\n".join(formula_lines)

    # Write to Summary
    ws_sum["A4"] = "Context"
    ws_sum["B4"] = context_agg
    ws_sum["A5"] = "Formulas"
    ws_sum["B5"] = formulas_agg
    # Add Prompt cell for LLM usage
    prompt_text = (
        "Du er en dansk Excel-ekspert. Brug FELTERNE nedenfor som input:\n\n"
        "[Context] = celle B4 (sammenfattet kontekst)\n"
        "[Formulas] = celle B5 (én linje pr. formelgruppe: Ark | Forekomster | Normaliseret | Eksempel | Lokation)\n\n"
        "Opgave:\n"
        "1) Læs Context for at forstå faglige begreber og terminologi.\n"
        "2) For hver formelgruppe i Formulas, lav en pædagogisk forklaring i en tabel med kolonnerne:\n"
        "   - Ark (som angivet)\n"
        "   - Normaliseret formel (ufortolket tekst)\n"
        "   - Hvad gør formlen overordnet (kort beskrivelse)\n"
        "   - Centrale funktioner/led (punktvis: IF, IFERROR, DATEDIF, ROUNDUP, PMT, XLOOKUP, SWITCH m.fl.)\n"
        "   - Referencer (forklar absolutte vs. relative referencer; fx $F$66 vs. B297)\n"
        "   - Forretningsregler (særlige regler der kan aflæses; fx 'hvis X så 0; ellers PMT…')\n"
        "   - Eksempel (forklar Eksempel-formlen i almindeligt sprog; brug semikolon i funktions-eksempler)\n"
        "3) Brug dansk, præcist og uden fyldord. Bevar funktionsnavne på engelsk, brug semikolon som separator.\n"
        "4) Lav én række per formelgruppe; sorter efter Ark og dernæst Normaliseret.\n"
        "5) Antag at Formulas og Context er korrekte; gæt ikke på manglende tal.\n\n"
        "Output: En markdown-tabel, ingen ekstra tekst før eller efter."
    )
    ws_sum["A6"] = "Prompt"
    ws_sum["B6"] = prompt_text
    ws_sum.column_dimensions['A'].width = 18
    ws_sum.column_dimensions['B'].width = 140

    # Save
    wb_out.save(out_path)
    return out_path


def main() -> None:
    parser = argparse.ArgumentParser(description="Forhåndsvis udtræk af formler fra Excel.")
    parser.add_argument("--path", help="Sti til .xlsx/.xlsm/.xltx/.xltm. Udelad for at vælge via GUI.")
    parser.add_argument("--max-examples", type=int, default=10, help="Maks antal eksempler pr. formel")
    # Allow both --similarity and --no-similarity; default follows DEFAULT_SIMILARITY
    sim_group = parser.add_mutually_exclusive_group()
    sim_group.add_argument("--similarity", dest="similarity", action="store_true", help="Vis lighedsgrupper (overstyrer standard)")
    sim_group.add_argument("--no-similarity", dest="similarity", action="store_false", help="Skjul lighedsgrupper (overstyrer standard)")
    parser.set_defaults(similarity=DEFAULT_SIMILARITY)
    parser.add_argument("--out", help="Gem rapport til fil (txt). Hvis udeladt, printes til skærm.")
    parser.add_argument("--export-xlsx", help="Gem organiseret Excel-rapport til denne sti (.xlsx)")
    parser.add_argument("--dest-dir", help="Destinationsmappe til standard-eksport (hvis udeladt, vælges via GUI)")
    args = parser.parse_args()

    xlsx_path: Optional[str] = args.path or pick_file_gui()
    if not xlsx_path:
        print("Ingen fil valgt.")
        sys.exit(1)
    if not os.path.exists(xlsx_path):
        print(f"Filen findes ikke: {xlsx_path}")
        sys.exit(1)

    mapping = collect_formulas(xlsx_path)
    report = summarize(mapping, args.max_examples)
    if args.similarity:
        report += "\n\n" + summarize_similarity(mapping, args.max_examples)

    # Decide default export destinations if none provided
    did_export_any = False
    default_folder = None
    if not args.export_xlsx or not args.out:
        # Ask for destination folder if not provided via --dest-dir
        default_folder = args.dest_dir or pick_destination_dir_gui()
        if not default_folder:
            # Fallback: same folder as input file
            default_folder = os.path.dirname(os.path.abspath(xlsx_path))

    # Create a subfolder to keep outputs organized
    from datetime import datetime
    base_name = os.path.splitext(os.path.basename(xlsx_path))[0]
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_folder = os.path.join(default_folder, f"{base_name}_preview_{timestamp}")
    try:
        os.makedirs(output_folder, exist_ok=True)
    except Exception:
        # If we cannot create, fallback to default folder directly
        output_folder = default_folder

    # Excel export path
    excel_out_path = args.export_xlsx or os.path.join(output_folder, "preview_report.xlsx")
    if not excel_out_path.lower().endswith('.xlsx'):
        excel_out_path = os.path.join(output_folder, "preview_report.xlsx")
    saved_xlsx = export_to_excel(xlsx_path, mapping, args.similarity, args.max_examples, excel_out_path)
    print(f"Excel-rapport gemt: {saved_xlsx}")
    did_export_any = True

    # Text export path
    text_out_path = args.out or os.path.join(output_folder, "preview_report.txt")
    with open(text_out_path, "w", encoding="utf-8") as f:
        f.write(report)
    print(f"Tekstrapport gemt: {text_out_path}")
    did_export_any = True

    # Also print to console for quick glance
    if not args.out:
        print(report)


if __name__ == "__main__":
    main()


